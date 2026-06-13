from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .utils import root_path


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center")
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(max(12, max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 50) + 1)) + 2), 55)
        ws.freeze_panes = "A2"
    wb.save(path)


def write_report(portfolio_rows: list[dict], forecasts: list, retrain_alerts: list[dict], settings: dict) -> Path:
    output = root_path(settings.get("reports", {}).get("output_file", "reports/bias_forecast_report.xlsx"))
    output.parent.mkdir(parents=True, exist_ok=True)
    forecast_rows = [asdict(f) for f in forecasts]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(portfolio_rows).to_excel(writer, sheet_name="Portfolio Forecast", index=False)
        pd.DataFrame([r for r in forecast_rows if r["model_type"] == "daily"]).to_excel(writer, sheet_name="Daily Forecasts", index=False)
        pd.DataFrame([r for r in forecast_rows if r["model_type"] == "weekly"]).to_excel(writer, sheet_name="Weekly Forecasts", index=False)
        pd.DataFrame(retrain_alerts).to_excel(writer, sheet_name="Retrain Alerts", index=False)
    _style_workbook(output)
    return output
